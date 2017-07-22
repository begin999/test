#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Thu Apr 20 21:15:12 2017

@author: catherine
"""

import openpyxl
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook


wb=openpyxl.load_workbook('example.xlsx')
wb.get_sheet_names()
sheet=wb.get_sheet_by_name('Sheet1')

tabs=np.array(['Sheet1', 'Sheet2'])
for tab in tabs:
    df=pd.read_excel('example.xlsx', sheetname=tab)
    df_left=df.iloc[:, 0:7]
    #print(df_left.shape)
    df_right=df.iloc[:, 9:16]
    print(df_right.shape)
    df_right.columns=['COB','Region','Name',	'Item','Quantity','Rate','Total']
    df_left.append(df_right)
    #print(df.shape)

df=pd.read_excel('example.xlsx', sheetname='Sheet1')
df_left=df.iloc[:, 0:7]
df_right=df.iloc[:, 9:16]
df_right.columns=['COB','Region','Name',	'Item','Quantity','Rate','Total']
df_new=df_left.append(df_right)
#df_new.to_excel('saved.xlsx', index=False)

writer=pd.ExcelWriter('example.xlsx')
df_new.to_excel(writer, 'sheeet',index=False)
#df_new.to_excel(writer, 'Sheet2', index=False)
writer.save()


book = load_workbook('example.xlsx')
writer = pd.ExcelWriter('example.xlsx', engine='openpyxl') 
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
if writer.sheets is 'Sheet1':
    df=book.copy_worksheet('Sheet1')
    df_left=df.iloc[:, 0:7]
    df_right=df.iloc[:, 8:15]
    #print(df_right.shape)
    df_right.columns=['COB','Region','Name',	'Item','Quantity','Rate','Total']
    df=df_left.append(df_right)
writer.save()



